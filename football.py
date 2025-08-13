import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, LineChart, Reference
from collections import defaultdict

def style_excel_sheet(worksheet):
    #Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    color1_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    color2_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    if worksheet.max_row == 0:
        return 

    #Header Style
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    #Rows Styele
    for index, row in enumerate(worksheet.iter_rows(min_row=2)):
        fill = color1_fill if index % 2 == 0 else color2_fill
        for cell in row:
            cell.fill = fill

    # Adjust column widths for better readability
    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_raw_data_sheet(workbook, csv_file_path):
    """
    Reads data from a CSV file and creates a styled sheet in the workbook.
    """
    print("Creating 'Raw Football Results' sheet...")
    worksheet = workbook.create_sheet(title="Raw Football Results", index=0)

    with open(csv_file_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            worksheet.append(row)
    
    style_excel_sheet(worksheet)
    print("Successfully created and styled the raw data sheet.")
    # Return the header for other functions to use
    with open(csv_file_path, 'r', encoding='utf-8') as f:
        return next(csv.reader(f))


def create_yearly_analysis_sheet(workbook, csv_file_path, header):
    """
    Analyzes football data by year and creates a sheet with data and charts.
    This function does NOT use pandas.
    """
    print("Analyzing data for 'Yearly Analysis' sheet...")
    worksheet = workbook.create_sheet(title="Yearly Analysis", index=1)
    
    # Find column indices from the header
    try:
        season_idx = header.index('Season')
        fthg_idx = header.index('FTHG') # Full Time Home Goals
        ftag_idx = header.index('FTAG') # Full Time Away Goals
        ftr_idx = header.index('FTR')   # Full Time Result
    except ValueError as e:
        print(f"Error: Missing expected column in CSV: {e}")
        return

    yearly_data = defaultdict(lambda: {'TotalGoals': 0, 'HomeWins': 0, 'AwayWins': 0})

    with open(csv_file_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader) # Skip header row
        for row in reader:
            try:
                season = row[season_idx]
                yearly_data[season]['TotalGoals'] += int(row[fthg_idx]) + int(row[ftag_idx])
                if row[ftr_idx] == 'H':
                    yearly_data[season]['HomeWins'] += 1
                elif row[ftr_idx] == 'A':
                    yearly_data[season]['AwayWins'] += 1
            except (ValueError, IndexError):
                # Skip rows with conversion errors or incorrect number of columns
                continue
    
    # Write aggregated data to the sheet
    worksheet.append(['Season', 'Total Goals', 'Home Wins', 'Away Wins', 'Win Difference'])
    sorted_seasons = sorted(yearly_data.keys())
    
    for season in sorted_seasons:
        data = yearly_data[season]
        win_diff = data['HomeWins'] - data['AwayWins']
        worksheet.append([season, data['TotalGoals'], data['HomeWins'], data['AwayWins'], win_diff])

    style_excel_sheet(worksheet)
    
    data_rows = len(sorted_seasons) + 1

    # --- Create Bar Chart for Total Goals per Year ---
    bar_chart = BarChart()
    bar_chart.title = "Total Goals Scored per Season"
    bar_chart.y_axis.title = "Goals"
    bar_chart.x_axis.title = "Season"
    data = Reference(worksheet, min_col=2, min_row=1, max_row=data_rows, max_col=2)
    cats = Reference(worksheet, min_col=1, min_row=2, max_row=data_rows)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    worksheet.add_chart(bar_chart, "G2")

    # --- Create Line Chart for Home vs. Away Win Difference ---
    line_chart = LineChart()
    line_chart.title = "Home Win vs Away Win Difference per Season"
    line_chart.y_axis.title = "Difference (Home Wins - Away Wins)"
    line_chart.x_axis.title = "Season"
    data = Reference(worksheet, min_col=5, min_row=1, max_row=data_rows, max_col=5)
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(cats)
    worksheet.add_chart(line_chart, "G18")

    print("Successfully created the yearly analysis sheet with charts.")


def create_team_analysis_sheet(workbook, csv_file_path, header):
    """
    Calculates derived stats like win rates for each team and creates a new sheet.
    This function does NOT use pandas.
    """
    print("Performing team-level analysis for 'Team Win Rates' sheet...")
    worksheet = workbook.create_sheet(title="Team Win Rates", index=2)

    try:
        home_team_idx = header.index('HomeTeam')
        away_team_idx = header.index('AwayTeam')
        ftr_idx = header.index('FTR')
    except ValueError as e:
        print(f"Error: Missing expected column in CSV: {e}")
        return

    team_stats = defaultdict(lambda: {'HGP': 0, 'HW': 0, 'AGP': 0, 'AW': 0})

    with open(csv_file_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader) # Skip header
        for row in reader:
            try:
                home_team = row[home_team_idx]
                away_team = row[away_team_idx]
                
                # Aggregate games played
                team_stats[home_team]['HGP'] += 1
                team_stats[away_team]['AGP'] += 1
                
                # Aggregate wins
                if row[ftr_idx] == 'H':
                    team_stats[home_team]['HW'] += 1
                elif row[ftr_idx] == 'A':
                    team_stats[away_team]['AW'] += 1
            except IndexError:
                continue

    # Write the calculated stats to the sheet
    worksheet.append(['Team', 'Home Games Played', 'Home Wins', 'Home Win Rate (%)', 
                       'Away Games Played', 'Away Wins', 'Away Win Rate (%)'])
    
    sorted_teams = sorted(team_stats.keys())

    for team in sorted_teams:
        stats = team_stats[team]
        hgp, hw = stats['HGP'], stats['HW']
        agp, aw = stats['AGP'], stats['AW']
        
        home_win_rate = (hw / hgp * 100) if hgp > 0 else 0
        away_win_rate = (aw / agp * 100) if agp > 0 else 0
        
        worksheet.append([
            team, hgp, hw, f"{home_win_rate:.2f}",
            agp, aw, f"{away_win_rate:.2f}"
        ])

    style_excel_sheet(worksheet)
    print("Successfully created the team analysis sheet.")


if __name__ == '__main__':
    input_csv_file = 'results.csv'
    final_excel_output = 'football_analysis_report.xlsx'

    try:
        # Create a new workbook that will hold all our sheets
        main_workbook = Workbook()
        # Remove the default sheet that is created with a new workbook
        if 'Sheet' in main_workbook.sheetnames:
            main_workbook.remove(main_workbook['Sheet'])

        # Task 1: Create the raw data sheet and get the header
        csv_header = create_raw_data_sheet(main_workbook, input_csv_file)
        
        # Task 2: Create the yearly analysis sheet with charts
        create_yearly_analysis_sheet(main_workbook, input_csv_file, csv_header)
        
        # Task 3: Create the team analysis sheet
        create_team_analysis_sheet(main_workbook, input_csv_file, csv_header)
        
        # Save the single workbook with all three sheets
        main_workbook.save(final_excel_output)
        
        print(f"\nAll tasks completed successfully! The final report is saved as '{final_excel_output}'")

    except FileNotFoundError:
        print(f"Error: The file '{input_csv_file}' was not found.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
